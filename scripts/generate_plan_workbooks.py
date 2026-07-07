#!/usr/bin/env python3
"""Generate plan-v2.json and Excel workbooks for original + syllabus-optimized plans."""

from __future__ import annotations

import json
import subprocess
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
MATERIAL_ROOT = Path.home() / "Desktop/系规"
PLAN_V1 = PUBLIC / "plan.json"
PLAN_V2 = PUBLIC / "plan-v2.json"
XLSX_V1 = MATERIAL_ROOT / "2026-学习计划-原版.xlsx"
XLSX_V2 = MATERIAL_ROOT / "2026-学习计划-考纲优化版.xlsx"

EXAM_DATE = "2026-10-24"
START_DATE = "2026-07-07"

REGISTRATION = {
    "id": "register",
    "title": "网上报名",
    "start": "2026-08-15",
    "end": "2026-09-05",
    "note": "中国计算机技术职业资格网（ruankao.org.cn）· 8月中旬至9月上旬，切勿错过",
}

# Same lesson mapping as generate_plan.py; focus aligned to 2024 syllabus + S/A/B tiers
V2_WEEKS = [
    ("W1", "导学", "输入期", date(2026, 7, 7), date(2026, 7, 13), [1, 2]),
    ("W2", "基础", "输入期", date(2026, 7, 14), date(2026, 7, 20), list(range(3, 11))),
    ("W3", "基础", "输入期", date(2026, 7, 21), date(2026, 7, 27), list(range(11, 14))),
    ("W4", "规划", "输入期", date(2026, 7, 28), date(2026, 8, 3), list(range(14, 22))),
    ("W5", "规划", "输入期", date(2026, 8, 4), date(2026, 8, 10), list(range(22, 30))),
    ("W6", "规划", "输入期", date(2026, 8, 11), date(2026, 8, 17), list(range(30, 38))),
    ("W7", "规划", "输入期", date(2026, 8, 18), date(2026, 8, 24), list(range(38, 43))),
    ("W8", "管理", "核心期", date(2026, 8, 25), date(2026, 8, 31), list(range(43, 49))),
    ("W9", "管理", "核心期", date(2026, 9, 1), date(2026, 9, 7), list(range(49, 52))),
    ("W10", "管理", "巩固期", date(2026, 9, 8), date(2026, 9, 14), list(range(52, 70))),
    ("W11", "专项", "巩固期", date(2026, 9, 15), date(2026, 9, 21), list(range(70, 76))),
    ("W12", "专项", "巩固期", date(2026, 9, 22), date(2026, 9, 28), list(range(76, 84))),
    ("W13", "专题", "输出期", date(2026, 9, 29), date(2026, 10, 5), [903, 904, 905]),
    ("W14", "专题", "输出期", date(2026, 10, 6), date(2026, 10, 12), [901, 902]),
    ("W15", "冲刺", "冲刺期", date(2026, 10, 13), date(2026, 10, 19), []),
    ("W16", "冲刺", "冲刺期", date(2026, 10, 20), date(2026, 10, 23), []),
]

V2_FOCUS = {
    "W1": "【基础篇】考试认知+第2版考纲导读+494h时间预算",
    "W2": "【A档】第1章 信息系统与信息技术发展（新一代信息技术）",
    "W3": "【A档】第2章 数字中国与数智化发展",
    "W4": "【S档★】第4章 信息系统规划（CSF/BSP/价值链·论文高频）",
    "W5": "【A档】第5-6章 应用系统规划+云资源规划",
    "W6": "【A档】第7-8章 网络+数据资源规划 · ⚠️ 报名窗口已开（8/15起）",
    "W7": "【A档】第9-10章 信息安全+云原生 · ⚠️ 务必完成网上报名（9/5前）",
    "W8": "【S档★】第11章 信息系统治理（COBIT/治理框架）",
    "W9": "【S档★】第12章 IT服务管理（ITIL/ITSS·案例+论文主战场）",
    "W10": "【A档】第13-17章 人员/规范/技术/工具/项目管理",
    "W11": "【B档】第18-20章 智慧城市/园区/数字乡村（案例背景）",
    "W12": "【B/C档】第21-24章 转型/制造/消费/法规（速读+真题）",
    "W13": "【论文】IT服务/治理素材积累+案例专题课（专03-05）",
    "W14": "【案例+论文】案例分析专题+综合串讲（专01-02）",
    "W15": "综合题刷题+二模+错题复盘（S/A档回炉）",
    "W16": "考前串讲+查缺补漏+调整状态",
}

V2_CASE = {
    "W1": "—",
    "W2": "—",
    "W3": "—",
    "W4": "规划方法论可入论文",
    "W5": "云资源/应用规划案例点",
    "W6": "⚠️ 报名",
    "W7": "⚠️ 报名确认",
    "W8": "治理类案例",
    "W9": "IT服务案例+论文提纲",
    "W10": "项目管理案例",
    "W11": "智慧城市/园区案例",
    "W12": "法规选择题+案例素材",
    "W13": "论文框架+专题课",
    "W14": "案例专练+一模",
    "W15": "案例+论文二模",
    "W16": "—",
}

STAGES = [
    ("输入期", "2026-07-07", "2026-08-24", "基础篇+方法篇（第1-10章）", "视频精学+教材+章节题"),
    ("核心期", "2026-08-25", "2026-09-07", "S档攻坚（第11-12章 治理+服务）", "精读+默写+案例/论文素材"),
    ("巩固期", "2026-09-08", "2026-09-28", "A/B档+能力篇（第13-24章）", "速记+真题+专题巩固"),
    ("输出期", "2026-09-29", "2026-10-12", "案例+论文+专题课", "专01-05+写作+一模"),
    ("冲刺期", "2026-10-13", "2026-10-23", "刷题+模考+复盘", "二模+错题+调整"),
]


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])


def fmt_range(start: str, end: str) -> str:
    s = date.fromisoformat(start)
    e = date.fromisoformat(end)
    return f"{s.month}/{s.day}–{e.month}/{e.day}"


def week_dates(start: date, end: date) -> list[date]:
    days: list[date] = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur = cur.fromordinal(cur.toordinal() + 1)
    return days


def distribute(lesson_nos: list[int], days: list[date]) -> dict[str, list[int]]:
    if not lesson_nos:
        return {}
    buckets = {d.isoformat(): [] for d in days}
    keys = list(buckets.keys())
    for i, no in enumerate(lesson_nos):
        buckets[keys[i % len(keys)]].append(no)
    return buckets


def build_v2_weeks(lessons: dict) -> list[dict]:
    weeks_out = []
    for wid, stage, phase, start, end, lesson_nos in V2_WEEKS:
        days = week_dates(start, end)
        schedule = distribute(lesson_nos, days)
        tasks = []
        for day in days:
            for no in schedule.get(day.isoformat(), []):
                key = str(no)
                m = lessons.get(key, {"title": f"第{no}节", "missing": True})
                tasks.append(
                    {
                        "id": f"{wid.lower()}-{no:02d}",
                        "type": "video",
                        "lessonNo": no,
                        "title": m.get("title", f"第{no}节"),
                        "scheduledDate": day.isoformat(),
                        "missing": m.get("missing", key not in lessons),
                    }
                )
        weeks_out.append(
            {
                "id": wid,
                "stage": stage,
                "phase": phase,
                "focus": V2_FOCUS[wid],
                "caseArrangement": V2_CASE[wid],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "tasks": tasks,
            }
        )
    return weeks_out


def write_workbook(path: Path, title: str, weeks: list[dict], stages: list[tuple]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "学习计划"

    header_font = Font(bold=True, size=12)
    sub_font = Font(size=10, color="666666")
    hdr_fill = PatternFill("solid", fgColor="E8F0FE")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"考试日期 {EXAM_DATE.replace('-', '/')} · 16 周备考 · 报名提醒 8/15–9/5"
    ws["A2"].font = sub_font

    row = 4
    ws.cell(row, 1, "学习阶段").font = header_font
    ws.cell(row, 2, "开始时间").font = header_font
    ws.cell(row, 3, "截止时间").font = header_font
    ws.cell(row, 4, "任务详情").font = header_font
    ws.cell(row, 5, "说明").font = header_font
    for c in range(1, 6):
        ws.cell(row, c).fill = hdr_fill
    row += 1
    for stage, start, end, detail, note in stages:
        ws.cell(row, 1, stage)
        ws.cell(row, 2, start)
        ws.cell(row, 3, end)
        ws.cell(row, 4, detail)
        ws.cell(row, 5, note)
        row += 1

    row += 1
    ws.cell(row, 1, "周次").font = header_font
    ws.cell(row, 2, "日期范围").font = header_font
    ws.cell(row, 3, "学习阶段").font = header_font
    ws.cell(row, 4, "本周核心内容").font = header_font
    ws.cell(row, 5, "案例/论文安排").font = header_font
    for c in range(1, 6):
        ws.cell(row, c).fill = hdr_fill
    row += 1

    for w in weeks:
        ws.cell(row, 1, w["id"])
        ws.cell(row, 2, fmt_range(w["start"], w["end"]))
        ws.cell(row, 3, w.get("phase", ""))
        ws.cell(row, 4, w.get("focus", ""))
        ws.cell(row, 5, w.get("caseArrangement", w.get("casePaper", "—")))
        for c in range(1, 6):
            ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 22

    wb.save(path)


def main() -> None:
    ensure_openpyxl()

    if not PLAN_V1.exists():
        raise SystemExit(f"Missing {PLAN_V1}; run pnpm gen:plan first")

    MATERIAL_ROOT.mkdir(parents=True, exist_ok=True)

    plan_v1 = json.loads(PLAN_V1.read_text(encoding="utf-8"))
    lessons = plan_v1.get("lessons", {})
    weeks_v1 = plan_v1.get("weeks", [])

    for w in weeks_v1:
        w.setdefault("caseArrangement", "—")

    weeks_v2 = build_v2_weeks(lessons)

    plan_v2 = {
        **{k: v for k, v in plan_v1.items() if k not in ("weeks", "version")},
        "planId": "syllabus-v2",
        "planName": "考纲优化版（2024大纲·S/A/B权重）",
        "version": 3,
        "milestones": [REGISTRATION],
        "weeks": weeks_v2,
    }
    PLAN_V2.write_text(json.dumps(plan_v2, ensure_ascii=False, indent=2), encoding="utf-8")

    write_workbook(
        XLSX_V1,
        "2026考期系统规划与管理师·学习计划表（原版·视频课表）",
        weeks_v1,
        STAGES,
    )
    write_workbook(
        XLSX_V2,
        "2026考期系统规划与管理师·学习计划表（考纲优化版）",
        weeks_v2,
        STAGES,
    )

    print(f"Wrote {PLAN_V2}")
    print(f"Wrote {XLSX_V1}")
    print(f"Wrote {XLSX_V2}")


if __name__ == "__main__":
    main()
